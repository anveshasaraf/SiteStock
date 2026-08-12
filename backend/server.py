"""BuildTrack API - re-platformed on Supabase Postgres + Auth."""
from dotenv import load_dotenv
from pathlib import Path
load_dotenv(Path(__file__).parent / ".env")

import io
import csv
import logging
import os
import uuid
from datetime import datetime, timezone, timedelta
from typing import Optional, List, Literal

import httpx
from fastapi import (
    FastAPI, APIRouter, HTTPException, Depends, Request,
    Response, Query, UploadFile, File,
)
from fastapi.responses import StreamingResponse
from fastapi.middleware.gzip import GZipMiddleware
from starlette.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field, EmailStr
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

from db import get_pool, close_pool
from auth_supabase import (
    get_current_user, require_super_admin, require_site_role, get_user_site_role,
)
from storage_supabase import (
    upload_file, get_signed_url, delete_file,
    make_storage_path, ALLOWED_MIME, MAX_BYTES,
)

# ── App setup ────────────────────────────────────────────────────────────────

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("buildtrack")

limiter = Limiter(key_func=get_remote_address)
app = FastAPI(title="BuildTrack API", docs_url="/api/docs", openapi_url="/api/openapi.json")
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
app.add_middleware(GZipMiddleware, minimum_size=500)

api = APIRouter(prefix="/api")

SUPABASE_URL  = os.environ.get("SUPABASE_URL", "")
SERVICE_KEY   = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

# ── Helpers ──────────────────────────────────────────────────────────────────

def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def new_id() -> str:
    return str(uuid.uuid4())


def inr(n) -> float:
    return round(float(n or 0), 2)


def _row_to_dict(row) -> dict:
    """Convert asyncpg Record to plain dict, stringifying UUIDs."""
    if row is None:
        return {}
    d = {}
    for k, v in dict(row).items():
        if isinstance(v, uuid.UUID):
            d[k] = str(v)
        elif isinstance(v, (datetime,)):
            d[k] = v.isoformat()
        else:
            d[k] = v
    return d


def _rows(records) -> List[dict]:
    return [_row_to_dict(r) for r in records]


# ── Pydantic Models ──────────────────────────────────────────────────────────

class SiteIn(BaseModel):
    name: str
    location: Optional[str] = ""
    code: Optional[str] = ""


class MembershipIn(BaseModel):
    user_id: str
    site_id: str
    role: Literal["viewer", "logger", "manager", "site_admin"]


class ProvisionUserIn(BaseModel):
    """Super-admin / site-admin creates a new user."""
    name: str
    email: Optional[EmailStr] = None
    phone: Optional[str] = None       # +91XXXXXXXXXX for OTP users
    role: Literal["viewer", "logger", "manager", "site_admin"] = "logger"
    site_id: str                       # initial site assignment
    is_super_admin: bool = False


class CategoryIn(BaseModel):
    name: str


class ItemIn(BaseModel):
    name: str
    category: str
    unit: str = "nos"
    min_stock: float = 0
    max_stock: float = 0
    rate: float = 0
    description: Optional[str] = ""


class SupplierIn(BaseModel):
    name: str
    contact: Optional[str] = ""
    phone: Optional[str] = ""
    address: Optional[str] = ""


class InvoiceLine(BaseModel):
    item_id: str
    item_name: str
    unit: str
    quantity: float
    rate: float
    amount: float


class InvoiceIn(BaseModel):
    invoice_number: str
    supplier_id: Optional[str] = None
    supplier_name: str
    invoice_date: str
    gst_percent: float = 0
    lines: List[InvoiceLine]
    notes: Optional[str] = ""
    attachment_path: Optional[str] = ""
    attachment_name: Optional[str] = ""


class MovementIn(BaseModel):
    item_id: str
    quantity: float
    rate: float = 0
    type: Literal["inward", "outward", "consumption"]
    reference: Optional[str] = ""
    notes: Optional[str] = ""
    issued_to: Optional[str] = ""


class PhysicalCountIn(BaseModel):
    item_id: str
    counted_qty: float
    notes: Optional[str] = ""
    adjust: bool = False
    photo_path: Optional[str] = ""
    photo_name: Optional[str] = ""


# ── Health ───────────────────────────────────────────────────────────────────

@api.get("/")
async def root():
    return {"service": "BuildTrack", "status": "ok", "version": "2.0"}


# ── Auth / Me ────────────────────────────────────────────────────────────────

@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    """Returns the current user's profile + their site memberships."""
    pool = await get_pool()
    async with pool.acquire() as conn:
        sites = await conn.fetch(
            """
            SELECT s.id, s.name, s.code, s.location, m.role
            FROM memberships m
            JOIN sites s ON s.id = m.site_id
            WHERE m.user_id = $1 AND s.deleted_at IS NULL
            ORDER BY s.name
            """,
            uuid.UUID(user["id"]),
        )
    return {
        **user,
        "sites": [
            {
                "id": str(r["id"]),
                "name": r["name"],
                "code": r["code"],
                "location": r["location"],
                "role": r["role"],
            }
            for r in sites
        ],
    }


# ── Org-admin: Sites ─────────────────────────────────────────────────────────

@api.get("/org/sites")
async def list_all_sites(user: dict = Depends(require_super_admin)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT id, name, code, location, created_at FROM sites WHERE deleted_at IS NULL ORDER BY name"
        )
    return _rows(rows)


@api.post("/org/sites")
async def create_site(data: SiteIn, user: dict = Depends(require_super_admin)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO sites (name, code, location, created_by)
            VALUES ($1, $2, $3, $4)
            RETURNING id, name, code, location, created_at
            """,
            data.name, data.code or "", data.location or "",
            uuid.UUID(user["id"]),
        )
    return _row_to_dict(row)


@api.put("/org/sites/{site_id}")
async def update_site(site_id: str, data: SiteIn, user: dict = Depends(require_site_role("site_admin"))):
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            UPDATE sites SET name=$1, code=$2, location=$3, updated_at=NOW()
            WHERE id=$4 AND deleted_at IS NULL
            RETURNING id, name, code, location
            """,
            data.name, data.code or "", data.location or "",
            uuid.UUID(site_id),
        )
    if not row:
        raise HTTPException(404, "Site not found")
    return _row_to_dict(row)


@api.delete("/org/sites/{site_id}")
async def delete_site(site_id: str, user: dict = Depends(require_super_admin)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            "UPDATE sites SET deleted_at=NOW() WHERE id=$1", uuid.UUID(site_id)
        )
    return {"ok": True}


# ── Org-admin: Users + Memberships ───────────────────────────────────────────

@api.get("/org/users")
async def list_users(_: dict = Depends(require_super_admin)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT id, name, email, phone, is_super_admin, created_at FROM profiles ORDER BY name"
        )
    return _rows(rows)


@api.post("/org/users")
@limiter.limit("20/hour")
async def provision_user(request: Request, data: ProvisionUserIn, admin: dict = Depends(require_super_admin)):
    """
    Creates a Supabase Auth user (email OR phone) + profile + initial membership.
    Uses Supabase Admin API so no email/SMS confirmation is required for provisioned accounts.
    """
    if not data.email and not data.phone:
        raise HTTPException(400, "Either email or phone is required")

    # 1. Create auth user via Supabase Admin API
    payload: dict = {"user_metadata": {"name": data.name}}
    if data.email:
        payload["email"] = data.email
        payload["password"] = new_id()[:16]   # random temporary password (user will reset)
        payload["email_confirm"] = True
    else:
        payload["phone"] = data.phone
        payload["phone_confirm"] = True

    async with httpx.AsyncClient(timeout=20) as client:
        resp = await client.post(
            f"{SUPABASE_URL}/auth/v1/admin/users",
            json=payload,
            headers={
                "apikey": SERVICE_KEY,
                "Authorization": f"Bearer {SERVICE_KEY}",
            },
        )
    if resp.status_code not in (200, 201):
        raise HTTPException(400, f"Failed to create auth user: {resp.json().get('message', resp.text)}")

    auth_user_id = resp.json()["id"]
    pool = await get_pool()
    async with pool.acquire() as conn:
        # Upsert profile (trigger may have already created it)
        await conn.execute(
            """
            INSERT INTO profiles (id, name, email, phone, is_super_admin)
            VALUES ($1, $2, $3, $4, $5)
            ON CONFLICT (id) DO UPDATE
            SET name=$2, email=$3, phone=$4, is_super_admin=$5, updated_at=NOW()
            """,
            uuid.UUID(auth_user_id), data.name,
            data.email, data.phone, data.is_super_admin,
        )
        # Initial site membership
        await conn.execute(
            """
            INSERT INTO memberships (user_id, site_id, role, created_by)
            VALUES ($1, $2, $3, $4)
            ON CONFLICT (user_id, site_id) DO UPDATE SET role=$3
            """,
            uuid.UUID(auth_user_id), uuid.UUID(data.site_id),
            data.role, uuid.UUID(admin["id"]),
        )

    return {"id": auth_user_id, "name": data.name, "email": data.email, "phone": data.phone}


@api.delete("/org/users/{user_id}")
async def delete_user(user_id: str, _: dict = Depends(require_super_admin)):
    # Delete from Supabase Auth (cascade to profiles via FK)
    async with httpx.AsyncClient(timeout=15) as client:
        await client.delete(
            f"{SUPABASE_URL}/auth/v1/admin/users/{user_id}",
            headers={"apikey": SERVICE_KEY, "Authorization": f"Bearer {SERVICE_KEY}"},
        )
    return {"ok": True}


@api.get("/org/memberships/{site_id}")
async def list_memberships(site_id: str, user: dict = Depends(require_site_role("site_admin"))):
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT m.id, m.user_id, m.role, p.name, p.email, p.phone
            FROM memberships m
            JOIN profiles p ON p.id = m.user_id
            WHERE m.site_id = $1
            ORDER BY p.name
            """,
            uuid.UUID(site_id),
        )
    return _rows(rows)


@api.post("/org/memberships")
async def add_membership(data: MembershipIn, admin: dict = Depends(get_current_user)):
    if not admin.get("is_super_admin") and get_user_site_role(admin, data.site_id) != "site_admin":
        raise HTTPException(403, "site_admin or super_admin required")
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            """
            INSERT INTO memberships (user_id, site_id, role, created_by)
            VALUES ($1, $2, $3, $4)
            ON CONFLICT (user_id, site_id) DO UPDATE SET role=$3
            """,
            uuid.UUID(data.user_id), uuid.UUID(data.site_id),
            data.role, uuid.UUID(admin["id"]),
        )
    return {"ok": True}


@api.delete("/org/memberships/{membership_id}")
async def remove_membership(membership_id: str, admin: dict = Depends(get_current_user)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT site_id FROM memberships WHERE id=$1", uuid.UUID(membership_id)
        )
        if not row:
            raise HTTPException(404, "Membership not found")
        if not admin.get("is_super_admin") and get_user_site_role(admin, str(row["site_id"])) != "site_admin":
            raise HTTPException(403, "site_admin or super_admin required")
        await conn.execute("DELETE FROM memberships WHERE id=$1", uuid.UUID(membership_id))
    return {"ok": True}


# ── Site-level info ──────────────────────────────────────────────────────────

@api.get("/p/{site_id}/info")
async def site_info(site_id: str, user: dict = Depends(require_site_role("viewer"))):
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT id, name, code, location FROM sites WHERE id=$1 AND deleted_at IS NULL",
            uuid.UUID(site_id),
        )
    if not row:
        raise HTTPException(404, "Site not found")
    return {**_row_to_dict(row), "role": get_user_site_role(user, site_id)}


# ── Categories (global) ──────────────────────────────────────────────────────

@api.get("/categories")
async def list_categories(_: dict = Depends(get_current_user)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch("SELECT id, name FROM categories ORDER BY name")
    return _rows(rows)


@api.post("/categories")
async def create_category(data: CategoryIn, _: dict = Depends(get_current_user)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO categories (name) VALUES ($1)
            ON CONFLICT (name) DO UPDATE SET name=EXCLUDED.name
            RETURNING id, name
            """,
            data.name,
        )
    return _row_to_dict(row)


@api.delete("/categories/{cid}")
async def delete_category(cid: str, _: dict = Depends(require_super_admin)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("DELETE FROM categories WHERE id=$1", uuid.UUID(cid))
    return {"ok": True}


# ── Items (global master) ────────────────────────────────────────────────────

@api.get("/items")
async def list_items(
    _: dict = Depends(get_current_user),
    page: int = Query(1, ge=1),
    page_size: int = Query(200, ge=1, le=500),
    q: Optional[str] = None,
):
    offset = (page - 1) * page_size
    pool = await get_pool()
    async with pool.acquire() as conn:
        if q:
            rows = await conn.fetch(
                """
                SELECT id, name, category, unit, min_stock, max_stock, rate, description
                FROM items WHERE deleted_at IS NULL AND name ILIKE $1
                ORDER BY name LIMIT $2 OFFSET $3
                """,
                f"%{q}%", page_size, offset,
            )
        else:
            rows = await conn.fetch(
                """
                SELECT id, name, category, unit, min_stock, max_stock, rate, description
                FROM items WHERE deleted_at IS NULL
                ORDER BY name LIMIT $1 OFFSET $2
                """,
                page_size, offset,
            )
    return _rows(rows)


@api.post("/items")
async def create_item(data: ItemIn, user: dict = Depends(get_current_user)):
    _assert_can_write_master(user)
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO items (name, category, unit, min_stock, max_stock, rate, description)
            VALUES ($1,$2,$3,$4,$5,$6,$7)
            RETURNING id, name, category, unit, min_stock, max_stock, rate, description
            """,
            data.name, data.category, data.unit,
            data.min_stock, data.max_stock, data.rate, data.description or "",
        )
    return _row_to_dict(row)


@api.put("/items/{iid}")
async def update_item(iid: str, data: ItemIn, user: dict = Depends(get_current_user)):
    _assert_can_write_master(user)
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            UPDATE items SET name=$1, category=$2, unit=$3, min_stock=$4,
                max_stock=$5, rate=$6, description=$7, updated_at=NOW()
            WHERE id=$8 AND deleted_at IS NULL
            RETURNING id, name, category, unit, min_stock, max_stock, rate, description
            """,
            data.name, data.category, data.unit,
            data.min_stock, data.max_stock, data.rate, data.description or "",
            uuid.UUID(iid),
        )
    if not row:
        raise HTTPException(404, "Item not found")
    return _row_to_dict(row)


@api.delete("/items/{iid}")
async def delete_item(iid: str, _: dict = Depends(require_super_admin)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            "UPDATE items SET deleted_at=NOW() WHERE id=$1", uuid.UUID(iid)
        )
    return {"ok": True}


def _assert_can_write_master(user: dict):
    """Items/Suppliers master data: manager+ in any site OR super_admin."""
    if user.get("is_super_admin"):
        return
    if any(r in ("manager", "site_admin") for r in user["memberships"].values()):
        return
    raise HTTPException(403, "Manager role or higher required to modify master data")


# ── Suppliers (global) ───────────────────────────────────────────────────────

@api.get("/suppliers")
async def list_suppliers(_: dict = Depends(get_current_user)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT id, name, contact, phone, address FROM suppliers WHERE deleted_at IS NULL ORDER BY name"
        )
    return _rows(rows)


@api.post("/suppliers")
async def create_supplier(data: SupplierIn, user: dict = Depends(get_current_user)):
    _assert_can_write_master(user)
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "INSERT INTO suppliers (name, contact, phone, address) VALUES ($1,$2,$3,$4) RETURNING *",
            data.name, data.contact or "", data.phone or "", data.address or "",
        )
    return _row_to_dict(row)


@api.put("/suppliers/{sid}")
async def update_supplier(sid: str, data: SupplierIn, user: dict = Depends(get_current_user)):
    _assert_can_write_master(user)
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            UPDATE suppliers SET name=$1, contact=$2, phone=$3, address=$4, updated_at=NOW()
            WHERE id=$5 AND deleted_at IS NULL
            RETURNING id, name, contact, phone, address
            """,
            data.name, data.contact or "", data.phone or "", data.address or "",
            uuid.UUID(sid),
        )
    if not row:
        raise HTTPException(404, "Supplier not found")
    return _row_to_dict(row)


@api.delete("/suppliers/{sid}")
async def delete_supplier(sid: str, user: dict = Depends(get_current_user)):
    _assert_can_write_master(user)
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            "UPDATE suppliers SET deleted_at=NOW() WHERE id=$1", uuid.UUID(sid)
        )
    return {"ok": True}


# ── Invoices (site-scoped) ───────────────────────────────────────────────────

@api.get("/p/{site_id}/invoices")
async def list_invoices(
    site_id: str,
    user: dict = Depends(require_site_role("viewer")),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
):
    offset = (page - 1) * page_size
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT i.id, i.invoice_number, i.supplier_name, i.invoice_date,
                   i.gst_percent, i.subtotal, i.gst_amount, i.total,
                   i.notes, i.attachment_path, i.attachment_name, i.created_at,
                   json_agg(json_build_object(
                       'id', il.id, 'item_id', il.item_id, 'item_name', il.item_name,
                       'unit', il.unit, 'quantity', il.quantity, 'rate', il.rate, 'amount', il.amount
                   ) ORDER BY il.id) AS lines
            FROM invoices i
            LEFT JOIN invoice_lines il ON il.invoice_id = i.id
            WHERE i.site_id=$1 AND i.deleted_at IS NULL
            GROUP BY i.id
            ORDER BY i.invoice_date DESC, i.created_at DESC
            LIMIT $2 OFFSET $3
            """,
            uuid.UUID(site_id), page_size, offset,
        )
    result = []
    for r in rows:
        d = _row_to_dict(r)
        if isinstance(d.get("lines"), str):
            import json
            d["lines"] = json.loads(d["lines"])
        result.append(d)
    return result


@api.post("/p/{site_id}/invoices")
async def create_invoice(
    site_id: str, data: InvoiceIn,
    user: dict = Depends(require_site_role("logger")),
):
    subtotal = sum(line.amount for line in data.lines)
    gst_amount = round(subtotal * data.gst_percent / 100, 2)
    total = round(subtotal + gst_amount, 2)

    pool = await get_pool()
    async with pool.acquire() as conn:
        async with conn.transaction():
            # Insert invoice
            inv_row = await conn.fetchrow(
                """
                INSERT INTO invoices
                    (invoice_number, supplier_id, supplier_name, site_id,
                     invoice_date, gst_percent, subtotal, gst_amount, total,
                     notes, attachment_path, attachment_name, created_by)
                VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13)
                RETURNING id, invoice_number
                """,
                data.invoice_number,
                uuid.UUID(data.supplier_id) if data.supplier_id else None,
                data.supplier_name, uuid.UUID(site_id),
                data.invoice_date, data.gst_percent,
                subtotal, gst_amount, total,
                data.notes or "", data.attachment_path or "", data.attachment_name or "",
                uuid.UUID(user["id"]),
            )
            inv_id = inv_row["id"]

            # Insert lines + create inward movements (in same transaction)
            for line in data.lines:
                await conn.execute(
                    """
                    INSERT INTO invoice_lines (invoice_id, item_id, item_name, unit, quantity, rate, amount)
                    VALUES ($1,$2,$3,$4,$5,$6,$7)
                    """,
                    inv_id,
                    uuid.UUID(line.item_id), line.item_name, line.unit,
                    line.quantity, line.rate, line.amount,
                )
                # Auto inward movement per line
                await conn.execute(
                    """
                    INSERT INTO movements
                        (item_id, item_name, site_id, quantity, rate, amount,
                         type, reference, notes, created_by)
                    VALUES ($1,$2,$3,$4,$5,$6,'inward',$7,$8,$9)
                    """,
                    uuid.UUID(line.item_id), line.item_name, uuid.UUID(site_id),
                    line.quantity, line.rate, line.amount,
                    f"INV:{data.invoice_number}",
                    f"Auto-created from invoice {data.invoice_number}",
                    uuid.UUID(user["id"]),
                )

    return {"id": str(inv_id), "invoice_number": data.invoice_number, "total": total}


@api.delete("/p/{site_id}/invoices/{inv_id}")
async def delete_invoice(
    site_id: str, inv_id: str,
    user: dict = Depends(require_site_role("manager")),
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        inv = await conn.fetchrow(
            "SELECT invoice_number FROM invoices WHERE id=$1 AND site_id=$2 AND deleted_at IS NULL",
            uuid.UUID(inv_id), uuid.UUID(site_id),
        )
        if not inv:
            raise HTTPException(404, "Invoice not found")
        async with conn.transaction():
            # Soft-delete the auto-inward movements for this invoice
            await conn.execute(
                "UPDATE movements SET deleted_at=NOW() WHERE site_id=$1 AND reference=$2",
                uuid.UUID(site_id), f"INV:{inv['invoice_number']}",
            )
            await conn.execute(
                "UPDATE invoices SET deleted_at=NOW() WHERE id=$1", uuid.UUID(inv_id)
            )
    return {"ok": True}


# ── Movements (site-scoped) ──────────────────────────────────────────────────

@api.get("/p/{site_id}/movements")
async def list_movements(
    site_id: str,
    user: dict = Depends(require_site_role("viewer")),
    mtype: Optional[str] = Query(None, alias="type"),
    days: Optional[int] = Query(None, ge=1, le=3650),
    page: int = Query(1, ge=1),
    page_size: int = Query(200, ge=1, le=500),
):
    offset = (page - 1) * page_size
    pool = await get_pool()
    async with pool.acquire() as conn:
        sql = """
            SELECT id, item_id, item_name, quantity, rate, amount, type,
                   reference, notes, issued_to, created_at
            FROM movements WHERE site_id=$1 AND deleted_at IS NULL
        """
        params: list = [uuid.UUID(site_id)]
        idx = 2
        if mtype:
            sql += f" AND type=${idx}::movement_type"; params.append(mtype); idx += 1
        if days:
            sql += f" AND created_at >= NOW() - INTERVAL '1 day' * ${idx}"; params.append(days); idx += 1
        sql += f" ORDER BY created_at DESC LIMIT ${idx} OFFSET ${idx+1}"
        params.extend([page_size, offset])
        rows = await conn.fetch(sql, *params)
    return _rows(rows)


@api.post("/p/{site_id}/movements")
async def create_movement(
    site_id: str, data: MovementIn,
    user: dict = Depends(require_site_role("logger")),
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        item = await conn.fetchrow(
            "SELECT id, name, rate FROM items WHERE id=$1 AND deleted_at IS NULL",
            uuid.UUID(data.item_id),
        )
        if not item:
            raise HTTPException(404, "Item not found")
        rate = data.rate or float(item["rate"])
        amount = round(rate * data.quantity, 2)
        row = await conn.fetchrow(
            """
            INSERT INTO movements
                (item_id, item_name, site_id, quantity, rate, amount,
                 type, reference, notes, issued_to, created_by)
            VALUES ($1,$2,$3,$4,$5,$6,$7::movement_type,$8,$9,$10,$11)
            RETURNING id, item_name, quantity, rate, amount, type, created_at
            """,
            uuid.UUID(data.item_id), item["name"], uuid.UUID(site_id),
            data.quantity, rate, amount,
            data.type, data.reference or "", data.notes or "", data.issued_to or "",
            uuid.UUID(user["id"]),
        )
    return _row_to_dict(row)


@api.delete("/p/{site_id}/movements/{mid}")
async def delete_movement(
    site_id: str, mid: str,
    user: dict = Depends(require_site_role("manager")),
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        res = await conn.execute(
            "UPDATE movements SET deleted_at=NOW() WHERE id=$1 AND site_id=$2 AND deleted_at IS NULL",
            uuid.UUID(mid), uuid.UUID(site_id),
        )
    if res == "UPDATE 0":
        raise HTTPException(404, "Movement not found")
    return {"ok": True}


# ── Stock Summary (period-filtered inward/outward/consumed + closing balance) ─

@api.get("/p/{site_id}/stock-summary")
async def stock_summary(
    site_id: str,
    user: dict = Depends(require_site_role("viewer")),
    days: int = Query(30, ge=1, le=3650),
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT
                i.id AS item_id,
                i.name AS item_name,
                i.unit,
                COALESCE(SUM(m.quantity) FILTER (WHERE m.type='inward'),      0) AS inward_qty,
                COALESCE(SUM(m.amount)   FILTER (WHERE m.type='inward'),      0) AS inward_value,
                COALESCE(SUM(m.quantity) FILTER (WHERE m.type='outward'),     0) AS outward_qty,
                COALESCE(SUM(m.amount)   FILTER (WHERE m.type='outward'),     0) AS outward_value,
                COALESCE(SUM(m.quantity) FILTER (WHERE m.type='consumption'), 0) AS consumed_qty,
                COALESCE(SUM(m.amount)   FILTER (WHERE m.type='consumption'), 0) AS consumed_value,
                COALESCE(sr.stock, 0)  AS closing_qty,
                COALESCE(sr.value, 0)  AS closing_value,
                COALESCE(sr.status, 'OK') AS status
            FROM items i
            LEFT JOIN movements m
                ON m.item_id = i.id
               AND m.site_id = $1
               AND m.deleted_at IS NULL
               AND m.created_at >= NOW() - (INTERVAL '1 day' * $2)
            LEFT JOIN stock_register sr ON sr.item_id = i.id AND sr.site_id = $1
            WHERE i.deleted_at IS NULL
              AND (m.id IS NOT NULL OR sr.item_id IS NOT NULL)
            GROUP BY i.id, i.name, i.unit, sr.stock, sr.value, sr.status
            ORDER BY i.name
            """,
            uuid.UUID(site_id), days,
        )
    return [
        {
            "item_id":       str(r["item_id"]),
            "item_name":     r["item_name"],
            "unit":          r["unit"],
            "inward_qty":    float(r["inward_qty"]),
            "inward_value":  float(r["inward_value"]),
            "outward_qty":   float(r["outward_qty"]),
            "outward_value": float(r["outward_value"]),
            "consumed_qty":  float(r["consumed_qty"]),
            "consumed_value":float(r["consumed_value"]),
            "closing_qty":   float(r["closing_qty"]),
            "closing_value": float(r["closing_value"]),
            "status":        r["status"],
        }
        for r in rows
    ]


# ── Stock Register (site-scoped) ─────────────────────────────────────────────

@api.get("/p/{site_id}/stock")
async def get_stock(
    site_id: str,
    user: dict = Depends(require_site_role("viewer")),
    status_filter: Optional[str] = Query(None, alias="status"),
    q: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(200, ge=1, le=500),
):
    offset = (page - 1) * page_size
    conditions = ["site_id=$1"]
    params: list = [uuid.UUID(site_id)]

    if status_filter:
        conditions.append(f"status=${len(params)+1}")
        params.append(status_filter.upper())
    if q:
        conditions.append(f"item_name ILIKE ${len(params)+1}")
        params.append(f"%{q}%")

    where = " AND ".join(conditions)
    params += [page_size, offset]
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            f"""
            SELECT site_id, site_name, item_id, item_name, category, unit,
                   inward, outward, consumption, stock, min_stock, max_stock,
                   auto_min_stock, rate, value, status
            FROM stock_register
            WHERE {where}
            ORDER BY item_name
            LIMIT ${len(params)-1} OFFSET ${len(params)}
            """,
            *params,
        )
    return _rows(rows)


# ── Physical Stock Audit (site-scoped) ───────────────────────────────────────

@api.get("/p/{site_id}/physical-stock")
async def list_physical(
    site_id: str,
    user: dict = Depends(require_site_role("viewer")),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=500),
):
    offset = (page - 1) * page_size
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT id, item_id, item_name, unit, counted_qty, system_qty, variance,
                   notes, photo_path, photo_name, adjusted, counted_by_name, created_at
            FROM physical_counts
            WHERE site_id=$1
            ORDER BY created_at DESC LIMIT $2 OFFSET $3
            """,
            uuid.UUID(site_id), page_size, offset,
        )
    return _rows(rows)


@api.post("/p/{site_id}/physical-stock")
async def create_physical(
    site_id: str, data: PhysicalCountIn,
    user: dict = Depends(require_site_role("logger")),
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        item = await conn.fetchrow(
            "SELECT id, name, unit, rate FROM items WHERE id=$1 AND deleted_at IS NULL",
            uuid.UUID(data.item_id),
        )
        if not item:
            raise HTTPException(404, "Item not found")

        # Compute system stock via SQL (replaces _system_stock Python function)
        system_row = await conn.fetchrow(
            """
            SELECT
                SUM(CASE WHEN type='inward' THEN quantity ELSE 0 END)
                - SUM(CASE WHEN type IN ('outward','consumption') THEN quantity ELSE 0 END) AS stock
            FROM movements
            WHERE item_id=$1 AND site_id=$2 AND deleted_at IS NULL
            """,
            uuid.UUID(data.item_id), uuid.UUID(site_id),
        )
        system_qty = float(system_row["stock"] or 0)
        variance = round(data.counted_qty - system_qty, 3)

        async with conn.transaction():
            count_row = await conn.fetchrow(
                """
                INSERT INTO physical_counts
                    (item_id, item_name, unit, site_id, counted_qty, system_qty,
                     variance, notes, photo_path, photo_name, adjusted,
                     counted_by, counted_by_name)
                VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13)
                RETURNING id, item_name, counted_qty, system_qty, variance, adjusted
                """,
                uuid.UUID(data.item_id), item["name"], item["unit"] or "",
                uuid.UUID(site_id), data.counted_qty, system_qty,
                variance, data.notes or "",
                data.photo_path or "", data.photo_name or "",
                bool(data.adjust and variance != 0),
                uuid.UUID(user["id"]), user.get("name", ""),
            )

            if data.adjust and variance != 0:
                mv_type = "inward" if variance > 0 else "outward"
                rate = float(item["rate"] or 0)
                count_id = str(count_row["id"])[:8]
                await conn.execute(
                    """
                    INSERT INTO movements
                        (item_id, item_name, site_id, quantity, rate, amount,
                         type, reference, notes, created_by)
                    VALUES ($1,$2,$3,$4,$5,$6,$7::movement_type,$8,$9,$10)
                    """,
                    uuid.UUID(data.item_id), item["name"], uuid.UUID(site_id),
                    abs(variance), rate, round(abs(variance) * rate, 2),
                    mv_type, f"ADJ:{count_id}",
                    f"Audit adjustment ({'+' if variance > 0 else ''}{variance})",
                    uuid.UUID(user["id"]),
                )

    return _row_to_dict(count_row)


# ── Dashboard (site-scoped) ──────────────────────────────────────────────────

@api.get("/p/{site_id}/dashboard")
async def dashboard(site_id: str, user: dict = Depends(require_site_role("viewer"))):
    pool = await get_pool()
    async with pool.acquire() as conn:
        # Stock summary
        stock_summary = await conn.fetchrow(
            """
            SELECT
                SUM(value)                          AS total_stock_value,
                COUNT(*) FILTER (WHERE status IN ('LOW','OUT')) AS low_count,
                COUNT(*) FILTER (WHERE status = 'HIGH')         AS high_count,
                COUNT(DISTINCT item_id)                         AS item_count
            FROM stock_register WHERE site_id=$1
            """,
            uuid.UUID(site_id),
        )

        # Purchase value (invoices)
        purchase = await conn.fetchval(
            "SELECT COALESCE(SUM(total), 0) FROM invoices WHERE site_id=$1 AND deleted_at IS NULL",
            uuid.UUID(site_id),
        )

        # Consumption value
        consumption_val = await conn.fetchval(
            """
            SELECT COALESCE(SUM(amount), 0) FROM movements
            WHERE site_id=$1 AND type IN ('consumption','outward') AND deleted_at IS NULL
            """,
            uuid.UUID(site_id),
        )

        # Low + High stock items (top 20)
        low_items = await conn.fetch(
            """
            SELECT item_name, category, unit, stock, status
            FROM stock_register WHERE site_id=$1 AND status IN ('LOW','OUT')
            ORDER BY stock LIMIT 20
            """,
            uuid.UUID(site_id),
        )
        high_items = await conn.fetch(
            """
            SELECT item_name, category, unit, stock, status
            FROM stock_register WHERE site_id=$1 AND status='HIGH'
            ORDER BY stock DESC LIMIT 20
            """,
            uuid.UUID(site_id),
        )

        # 30-day trend
        cutoff = datetime.now(timezone.utc) - timedelta(days=30)
        trend_rows = await conn.fetch(
            """
            SELECT
                created_at::date                                     AS day,
                SUM(CASE WHEN type='inward'      THEN quantity ELSE 0 END) AS inward,
                SUM(CASE WHEN type='outward'     THEN quantity ELSE 0 END) AS outward,
                SUM(CASE WHEN type='consumption' THEN quantity ELSE 0 END) AS consumption
            FROM movements
            WHERE site_id=$1 AND created_at >= $2 AND deleted_at IS NULL
            GROUP BY created_at::date
            ORDER BY day
            """,
            uuid.UUID(site_id), cutoff,
        )

    return {
        "total_purchase_value":     inr(purchase),
        "total_consumption_value":  inr(consumption_val),
        "total_stock_value":        inr(stock_summary["total_stock_value"]),
        "items_count":              stock_summary["item_count"] or 0,
        "low_stock_count":          stock_summary["low_count"] or 0,
        "high_stock_count":         stock_summary["high_count"] or 0,
        "low_stock":                _rows(low_items),
        "high_stock":               _rows(high_items),
        "trend":                    [
            {
                "day": str(r["day"]),
                "inward": float(r["inward"] or 0),
                "outward": float(r["outward"] or 0),
                "consumption": float(r["consumption"] or 0),
            }
            for r in trend_rows
        ],
    }


# ── Super-admin dashboard ────────────────────────────────────────────────────

@api.get("/org/dashboard")
async def org_dashboard(_: dict = Depends(require_super_admin)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        by_site = await conn.fetch(
            """
            SELECT site_name,
                   SUM(value)                                    AS stock_value,
                   COUNT(*) FILTER (WHERE status IN ('LOW','OUT')) AS low_count
            FROM stock_register
            GROUP BY site_name ORDER BY site_name
            """
        )
        totals = await conn.fetchrow(
            """
            SELECT COALESCE(SUM(value), 0)                           AS stock_value,
                   COUNT(*) FILTER (WHERE status IN ('LOW','OUT'))   AS low_count,
                   COUNT(*) FILTER (WHERE status = 'HIGH')           AS high_count
            FROM stock_register
            """
        )
    return {
        "total_stock_value":  inr(totals["stock_value"]),
        "total_low_count":    totals["low_count"] or 0,
        "total_high_count":   totals["high_count"] or 0,
        "by_site":            _rows(by_site),
    }


# ── CSV / XLSX Exports ───────────────────────────────────────────────────────

def _csv_response(rows: List[dict], filename: str, headers: List[str]) -> StreamingResponse:
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    w.writeheader()
    for r in rows:
        w.writerow(r)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@api.get("/p/{site_id}/export/stock")
async def export_stock(site_id: str, user: dict = Depends(require_site_role("viewer"))):
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT site_name, item_name, category, unit, inward, outward, consumption,
                   stock, min_stock, max_stock, rate, value, status
            FROM stock_register WHERE site_id=$1 ORDER BY item_name
            """,
            uuid.UUID(site_id),
        )
    return _csv_response(
        _rows(rows), "stock_register.csv",
        ["site_name","item_name","category","unit","inward","outward","consumption",
         "stock","min_stock","max_stock","rate","value","status"],
    )


@api.get("/p/{site_id}/export/stock.xlsx")
async def export_stock_xlsx(site_id: str, user: dict = Depends(require_site_role("viewer"))):
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT site_name, item_name, category, unit, inward, outward, consumption,
                   stock, min_stock, max_stock, rate, value, status
            FROM stock_register WHERE site_id=$1 ORDER BY item_name
            """,
            uuid.UUID(site_id),
        )
    headers = ["site_name","item_name","category","unit","inward","outward","consumption",
               "stock","min_stock","max_stock","rate","value","status"]
    widths   = [18, 28, 18, 8, 10, 10, 12, 10, 10, 10, 10, 14, 10]
    return _xlsx_response(_rows(rows), headers, "stock_register.xlsx", widths, "status")


@api.get("/p/{site_id}/export/invoices.xlsx")
async def export_invoices_xlsx(site_id: str, user: dict = Depends(require_site_role("viewer"))):
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT i.invoice_number, i.invoice_date, i.supplier_name,
                   il.item_name, il.quantity, il.unit, il.rate, il.amount,
                   i.gst_percent, i.total
            FROM invoices i
            JOIN invoice_lines il ON il.invoice_id = i.id
            WHERE i.site_id=$1 AND i.deleted_at IS NULL
            ORDER BY i.invoice_date DESC, i.invoice_number
            """,
            uuid.UUID(site_id),
        )
    headers = ["invoice_number","invoice_date","supplier_name","item_name",
               "quantity","unit","rate","amount","gst_percent","total"]
    widths  = [16, 14, 22, 26, 10, 8, 12, 14, 12, 14]
    return _xlsx_response(_rows(rows), headers, "invoices.xlsx", widths)


@api.get("/p/{site_id}/export/movements.xlsx")
async def export_movements_xlsx(
    site_id: str,
    user: dict = Depends(require_site_role("viewer")),
    mtype: Optional[str] = Query(None, alias="type"),
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        if mtype:
            rows = await conn.fetch(
                """
                SELECT created_at, type, item_name, quantity, rate, amount,
                       reference, issued_to, notes
                FROM movements
                WHERE site_id=$1 AND type=$2::movement_type AND deleted_at IS NULL
                ORDER BY created_at DESC
                """,
                uuid.UUID(site_id), mtype,
            )
        else:
            rows = await conn.fetch(
                """
                SELECT created_at, type, item_name, quantity, rate, amount,
                       reference, issued_to, notes
                FROM movements WHERE site_id=$1 AND deleted_at IS NULL
                ORDER BY created_at DESC
                """,
                uuid.UUID(site_id),
            )
    headers = ["created_at","type","item_name","quantity","rate","amount",
               "reference","issued_to","notes"]
    widths  = [22, 14, 24, 10, 12, 14, 18, 18, 28]
    return _xlsx_response(_rows(rows), headers, f"{mtype or 'movements'}.xlsx", widths)


# ── XLSX helper ──────────────────────────────────────────────────────────────

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def _xlsx_response(rows, headers, filename, widths=None, fill_key=None):
    wb = Workbook()
    ws = wb.active
    ws.title = filename.replace(".xlsx", "")[:31]
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="111827")
    fills = {
        "OUT":  PatternFill("solid", fgColor="FEE2E2"),
        "LOW":  PatternFill("solid", fgColor="FEF3C7"),
        "HIGH": PatternFill("solid", fgColor="DBEAFE"),
    }
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=i)
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
        if fill_key and r.get(fill_key) in fills:
            fill = fills[r[fill_key]]
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill
    for i, w in enumerate(widths or [18] * len(headers), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = ws.dimensions
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── File Upload / Serve ───────────────────────────────────────────────────────

@api.post("/upload")
@limiter.limit("30/minute")
async def upload_file_endpoint(
    request: Request,
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
):
    ext = (file.filename or "bin").rsplit(".", 1)[-1].lower()
    if ext not in ALLOWED_MIME:
        raise HTTPException(400, f"Unsupported file type .{ext}. Allowed: {', '.join(ALLOWED_MIME)}")
    data = await file.read()
    if len(data) > MAX_BYTES:
        raise HTTPException(400, "File too large - 8 MB maximum")

    storage_path = make_storage_path(user["id"], ext)
    try:
        result = await upload_file(storage_path, data, ALLOWED_MIME[ext])
    except RuntimeError as e:
        logger.exception("Upload failed: %s", e)
        raise HTTPException(500, "File upload failed - storage unavailable")

    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            """
            INSERT INTO files (storage_path, original_filename, content_type, size, uploaded_by)
            VALUES ($1,$2,$3,$4,$5)
            """,
            result["path"], file.filename or "",
            ALLOWED_MIME[ext], result["size"], uuid.UUID(user["id"]),
        )
    return {"path": result["path"], "name": file.filename}


@api.get("/files/signed")
async def signed_file_url(
    path: str = Query(...),
    _: dict = Depends(get_current_user),
):
    """Returns a 30-minute signed URL for a private storage object."""
    pool = await get_pool()
    async with pool.acquire() as conn:
        rec = await conn.fetchrow(
            "SELECT id FROM files WHERE storage_path=$1 AND is_deleted=false", path
        )
    if not rec:
        raise HTTPException(404, "File not found")
    try:
        url = await get_signed_url(path)
    except RuntimeError as e:
        raise HTTPException(500, str(e))
    return {"url": url, "expires_in": 1800}


# ── Startup / Shutdown ────────────────────────────────────────────────────────

@app.on_event("startup")
async def startup():
    # Eagerly open the pool so first requests aren't slow
    await get_pool()
    logger.info("BuildTrack API v2 ready - Supabase Postgres backend")


@app.on_event("shutdown")
async def shutdown():
    await close_pool()


# ── Security middleware ───────────────────────────────────────────────────────

@app.middleware("http")
async def security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"]         = "DENY"
    response.headers["X-XSS-Protection"]        = "1; mode=block"
    response.headers["Referrer-Policy"]          = "strict-origin-when-cross-origin"
    # Only set HSTS in production (behind HTTPS)
    if os.environ.get("ENVIRONMENT") == "production":
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    return response


# ── CORS ─────────────────────────────────────────────────────────────────────

_origins = [o.strip() for o in os.environ.get("CORS_ORIGINS", "http://localhost:3000").split(",") if o.strip()]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type"],
    max_age=600,
)

app.include_router(api)
